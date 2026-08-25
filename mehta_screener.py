#!/usr/bin/env python3
"""
Mehta 3/3 NSE Top-750 Screener

Run:
    python mehta_screener.py

Outputs:
    output/Mehta_Screener_YYYY-MM-DD.xlsx

Notes:
- Market-cap ranking is built from the configured universe source.
- Price/history uses yfinance.
- Fundamental PAT data uses Screener.in pages where available.
- Sector comparison uses Yahoo Finance sector ETFs/index proxies configured in config.json.
- "Record PAT" is conservative: the script compares the latest TTM PAT with the
  historical annual PAT series it can retrieve. Exceptional-item exclusion is
  not perfectly machine-readable across every company; such rows are flagged.
"""

import json, logging, math, re, sys, time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
import requests
import yfinance as yf
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


BASE = Path(__file__).resolve().parent
CFG = json.loads((BASE / "config.json").read_text(encoding="utf-8"))
OUT = BASE / CFG["output"]["directory"]
OUT.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=getattr(logging, CFG["runtime"].get("log_level", "INFO")),
    format="%(asctime)s | %(levelname)s | %(message)s"
)
log = logging.getLogger("mehta")
for _name in ("yfinance", "yfinance.multi", "yfinance.utils"):
    _yl = logging.getLogger(_name)
    _yl.setLevel(logging.CRITICAL)
    _yl.propagate = False

HEADERS = {
    "User-Agent": CFG["data"]["user_agent"],
    "Accept-Language": "en-US,en;q=0.9",
}

session = requests.Session()
session.headers.update(HEADERS)


def clean_symbol(s):
    s = str(s).strip().upper()
    return s.replace(".NS", "").replace(".BO", "")


def yahoo_symbol(s):
    raw = str(s).strip().upper()
    # Yahoo index symbols start with ^ and must not get an .NS suffix.
    if raw.startswith("^"):
        return raw
    if raw.endswith(".NS") or raw.endswith(".BO"):
        return raw
    return clean_symbol(raw) + ".NS"


def get_json(url, params=None):
    r = session.get(url, params=params, timeout=CFG["runtime"]["http_timeout"])
    r.raise_for_status()
    return r.json()


def fetch_nse_universe():
    """
    Try configured NSE constituent sources in order.
    A local CSV can be supplied in config.json for maximum reproducibility.
    Required columns: Symbol, optionally Company Name, Industry.
    """
    local = CFG["universe"].get("local_csv", "")
    if local:
        p = BASE / local
        if p.exists():
            try:
                df = pd.read_csv(p)
                normalized = normalize_universe(df)
                if len(normalized) >= CFG["universe"].get("minimum_universe_size", 1):
                    return normalized
                log.warning(
                    "Local universe %s contains only %d valid symbols; ignoring it and trying remote sources.",
                    p.name, len(normalized)
                )
            except Exception as e:
                log.warning("Local universe %s could not be read: %s", p.name, e)

    # NSE NIFTY 500 CSV is used as the base pool; top-750 requires an expanded
    # exchange universe, so if an expanded local CSV is configured it is preferred.
    urls = CFG["universe"]["urls"]
    for url in urls:
        try:
            log.info("Downloading universe: %s", url)
            r = session.get(url, timeout=min(8, int(CFG["runtime"].get("http_timeout", 12))))
            r.raise_for_status()
            if url.lower().endswith(".csv"):
                from io import StringIO
                df = pd.read_csv(StringIO(r.text))
            else:
                tables = pd.read_html(r.text)
                df = max(tables, key=len)
            df = normalize_universe(df)
            if len(df) >= CFG["universe"]["minimum_universe_size"]:
                return df
            log.warning("Universe only has %s rows; trying next source.", len(df))
        except Exception as e:
            log.warning("Universe source failed: %s", e)

    raise RuntimeError(
        "Could not obtain a sufficiently large NSE universe. "
        "Download an NSE universe CSV and set universe.local_csv in config.json."
    )


def normalize_universe(df):
    cols = {str(c).strip().lower(): c for c in df.columns}
    sym_col = next((cols[k] for k in ["symbol", "ticker", "code"] if k in cols), None)
    if sym_col is None:
        raise ValueError("Universe file needs a Symbol/Ticker/Code column.")

    out = pd.DataFrame()
    out["Symbol"] = df[sym_col].map(clean_symbol)
    for target, candidates in {
        "Company": ["company name", "company", "name"],
        "Industry": ["industry", "industry name", "sector"],
    }.items():
        c = next((cols[k] for k in candidates if k in cols), None)
        out[target] = df[c].astype(str) if c else ""

    out = out.drop_duplicates("Symbol")
    out = out[out["Symbol"].str.match(r"^[A-Z0-9&.-]+$", na=False)]
    return out.reset_index(drop=True)


def yf_download(symbols, period="2y"):
    """Download price history in small batches so one Yahoo failure cannot
    collapse the entire request into yfinance's "No objects to concatenate".
    """
    tickers = [yahoo_symbol(s) for s in symbols if clean_symbol(s)]
    if not tickers:
        raise RuntimeError(
            "The NSE universe is empty. Check nse_universe.csv or the configured NSE sources."
        )

    frames = []
    batch_size = int(CFG["runtime"].get("yfinance_batch_size", 100))
    retries = int(CFG["runtime"].get("yfinance_retries", 2))
    timeout = int(CFG["runtime"].get("http_timeout", 12))

    for start in range(0, len(tickers), batch_size):
        batch = tickers[start:start + batch_size]
        last_error = None
        for attempt in range(1, retries + 1):
            try:
                log.info("Downloading prices %d-%d/%d (attempt %d)",
                         start + 1, min(start + len(batch), len(tickers)),
                         len(tickers), attempt)
                data = yf.download(
                    tickers=batch,
                    period=period,
                    interval="1d",
                    auto_adjust=False,
                    progress=False,
                    group_by="ticker",
                    threads=True,
                    timeout=timeout,
                )
                if data is not None and not data.empty:
                    frames.append(data)
                    last_error = None
                    break
                last_error = RuntimeError("Yahoo returned no price rows")
            except Exception as e:
                last_error = e
                time.sleep(min(2 * attempt, 6))
        if last_error is not None:
            log.warning("Price batch %d-%d failed after %d attempts: %s",
                        start + 1, start + len(batch), retries, last_error)

    if not frames:
        raise RuntimeError(
            "Yahoo Finance returned no price data for any ticker. "
            "Check your internet connection, Yahoo availability, or try again later."
        )

    # Concatenate batches along columns.
    return pd.concat(frames, axis=1)


def extract_close(data, symbol):
    y = yahoo_symbol(symbol)
    if isinstance(data.columns, pd.MultiIndex):
        if y not in data.columns.get_level_values(0):
            return pd.Series(dtype=float)
        x = data[y]
        return x["Close"].dropna()
    return data["Close"].dropna()


def safe_return(series, days=252):
    if len(series) < days + 1:
        return np.nan
    return float(series.iloc[-1] / series.iloc[-days-1] - 1)


def get_price_metrics(close):
    if close.empty:
        return {}
    current = float(close.iloc[-1])
    ath = float(close.max())
    ema200 = float(close.ewm(span=200, adjust=False).mean().iloc[-1])
    r52 = safe_return(close, 252)
    return {
        "Current Price": current,
        "ATH": ath,
        "ATH % From High": current / ath - 1,
        "At/Near ATH": current >= ath * CFG["rules"]["ath_tolerance"],
        "200 EMA": ema200,
        "Below 200 EMA": current < ema200,
        "52W Return": r52,
    }


def screener_pat(symbol):
    """
    Best-effort extraction from Screener.in.
    Returns historical annual PAT and latest TTM PAT if available.
    """
    url = f"https://www.screener.in/company/{clean_symbol(symbol)}/consolidated/"
    try:
        r = session.get(url, timeout=min(8, int(CFG["runtime"].get("http_timeout", 12))))
        if r.status_code != 200:
            url = f"https://www.screener.in/company/{clean_symbol(symbol)}/"
            r = session.get(url, timeout=CFG["runtime"]["http_timeout"])
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")

        section = None
        for h2 in soup.find_all(["h2", "h3"]):
            if "Profit & Loss" in h2.get_text(" ", strip=True):
                section = h2.find_parent("section")
                break
        if section is None:
            return {}

        table = section.find("table")
        if table is None:
            return {}

        headers = [x.get_text(" ", strip=True) for x in table.find_all("th")]
        rows = []
        for tr in table.find_all("tr"):
            cells = [x.get_text(" ", strip=True) for x in tr.find_all(["th", "td"])]
            if cells:
                rows.append(cells)

        if not rows:
            return {}

        pat_row = None
        for row in rows:
            label = row[0].lower()
            if label in {"net profit", "profit after tax", "net profit / loss"}:
                pat_row = row
                break
        if pat_row is None:
            return {}

        def num(x):
            x = str(x).replace(",", "").replace("₹", "").strip()
            if x in {"", "-", "—", "nan"}:
                return np.nan
            m = re.search(r"-?\d+(?:\.\d+)?", x)
            return float(m.group()) if m else np.nan

        values = [num(x) for x in pat_row[1:]]
        # Screener normally has annual columns followed by TTM in recent layouts.
        years = headers[1:] if len(headers) > 1 else []
        annual = {}
        for y, v in zip(years, values):
            if np.isfinite(v):
                annual[str(y)] = v

        numeric = [v for v in values if np.isfinite(v)]
        if not numeric:
            return {}

        ttm = numeric[-1]
        annual_values = numeric[:-1] if len(numeric) > 1 else numeric
        record = max(annual_values) if annual_values else ttm

        return {
            "TTM PAT": ttm,
            "Record PAT": record,
            "Record PAT?": bool(ttm >= record * CFG["rules"]["pat_record_tolerance"]),
            "PAT Data Source": "Screener.in",
            "Exceptional Adjustment": "NOT AUTOMATICALLY EXCLUDED",
        }
    except Exception as e:
        log.debug("PAT failed %s: %s", symbol, e)
        return {}


def sector_benchmark(industry):
    mapping = CFG["sector_benchmarks"]
    text = str(industry).lower()
    for key, ticker in mapping.items():
        if key.lower() in text:
            return ticker
    return CFG["default_sector_benchmark"]


def benchmark_returns():
    """Fetch optional benchmark returns quietly and independently."""
    tickers = [CFG["benchmarks"]["nifty500"], CFG["default_sector_benchmark"]]
    tickers += list(CFG["sector_benchmarks"].values())
    tickers = list(dict.fromkeys(tickers))
    out = {}
    timeout = int(CFG["runtime"].get("benchmark_timeout", 8))

    for t in tickers:
        yahoo_ticker = yahoo_symbol(t)
        try:
            log.info("Downloading benchmark %s", yahoo_ticker)
            raw = yf.download(
                tickers=yahoo_ticker, period="2y", interval="1d",
                auto_adjust=False, progress=False, group_by="ticker",
                threads=False, timeout=timeout,
            )
            if raw is None or raw.empty:
                log.warning("Benchmark %s unavailable; skipping", yahoo_ticker)
                continue
            close = extract_close(raw, t)
            value = safe_return(close, 252)
            if np.isfinite(value):
                out[t] = value
            else:
                log.warning("Benchmark %s has insufficient history; skipping", yahoo_ticker)
        except Exception:
            log.warning("Benchmark %s unavailable; skipping", yahoo_ticker)

    return out


def score_row(row):
    price = bool(row.get("At/Near ATH", False))
    pat = bool(row.get("Record PAT?", False))
    rs = bool(row.get("Beats Nifty500", False) and row.get("Beats Sector", False))

    score = int(price) + int(pat) + int(rs)
    if score == 3:
        action = "SUPER PERFORMER - BUY/ADD"
    elif score == 2:
        action = "PERFORMER - HOLD / DON'T ADD"
    elif score == 1:
        action = "UNDERPERFORMER - EXIT"
    else:
        action = "UNDERPERFORMER - EXIT"

    alt_exit = bool(row.get("Below 200 EMA", False) and not row.get("Beats Nifty500", False))

    failures = []
    if not price:
        failures.append("Price not at/near ATH")
    if not pat:
        failures.append("PAT not record/unknown")
    if not rs:
        failures.append("Relative strength failed")
    if alt_exit:
        failures.append("Below 200 EMA + under Nifty500")

    row["Score"] = f"{score}/3"
    row["Action"] = action
    row["Alternative 200EMA Exit"] = alt_exit
    row["Failure Reason"] = "; ".join(failures)
    return row


def process_stock(symbol, industry, company, prices, bench):
    close = extract_close(prices, symbol)
    pm = get_price_metrics(close)
    if not pm:
        return {"Symbol": symbol, "Company": company, "Industry": industry,
                "Data Status": "No price data"}

    sec = sector_benchmark(industry)
    sec_ret = bench.get(sec, np.nan)
    nifty_ret = bench.get(CFG["benchmarks"]["nifty500"], np.nan)

    row = {
        "Symbol": symbol,
        "Company": company,
        "Industry": industry,
        "Sector Benchmark": sec,
        **pm,
        "Nifty500 52W Return": nifty_ret,
        "Sector 52W Return": sec_ret,
        "Beats Nifty500": bool(np.isfinite(pm["52W Return"]) and np.isfinite(nifty_ret)
                                and pm["52W Return"] > nifty_ret),
        "Beats Sector": bool(np.isfinite(pm["52W Return"]) and np.isfinite(sec_ret)
                             and pm["52W Return"] > sec_ret),
        "Data Status": "Price OK",
    }
    # PAT is the slowest part (one web request per stock). Only fetch it for
    # stocks that can still reach at least 2/3 or 3/3. This preserves the score
    # while eliminating hundreds of unnecessary Screener.in requests.
    pre_score = int(row["At/Near ATH"]) + int(row["Beats Nifty500"] and row["Beats Sector"])
    if pre_score > 0:
        row.update(screener_pat(symbol))
    else:
        row["PAT Data Source"] = "Skipped (not a 3/3 candidate)"
    row.setdefault("Record PAT?", False)
    row = score_row(row)
    return row


def export_excel(df, path):
    sheets = {
        "ALL 750": df,
        "3-3 SUPER": df[df["Score"] == "3/3"],
        "2-3 HOLD": df[df["Score"] == "2/3"],
        "1-3": df[df["Score"] == "1/3"],
        "0-3 EXIT": df[df["Score"] == "0/3"],
    }

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, x in sheets.items():
            x.to_excel(writer, sheet_name=name[:31], index=False)

        summary = pd.DataFrame([
            ["Run date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Universe size", len(df)],
            ["3/3 Super Performers", int((df["Score"] == "3/3").sum())],
            ["2/3 Performers", int((df["Score"] == "2/3").sum())],
            ["1/3 Underperformers", int((df["Score"] == "1/3").sum())],
            ["0/3 Underperformers", int((df["Score"] == "0/3").sum())],
        ], columns=["Metric", "Value"])
        summary.to_excel(writer, sheet_name="SUMMARY", index=False)

    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            width = min(max(max(len(str(c.value or "")) for c in col) + 2, 10), 35)
            ws.column_dimensions[get_column_letter(col[0].column)].width = width
    wb.save(path)


def main():
    start = time.time()
    log.info("Starting Mehta 3/3 NSE Screener (FAST MODE)")

    universe = fetch_nse_universe()
    # If source has market cap, rank it. Otherwise retain source order.
    if "Market Cap" in universe.columns:
        universe = universe.sort_values("Market Cap", ascending=False)
    universe = universe.head(int(CFG["universe"]["top_n"])).copy()

    symbols = universe["Symbol"].tolist()
    log.info("Screening %d symbols", len(symbols))

    prices = yf_download(symbols, period="2y")
    bench = benchmark_returns()

    results = []
    workers = int(CFG["runtime"].get("workers", 16))
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futs = {
            ex.submit(
                process_stock,
                r["Symbol"], r["Industry"], r["Company"], prices, bench
            ): r["Symbol"]
            for _, r in universe.iterrows()
        }
        for i, fut in enumerate(as_completed(futs), 1):
            try:
                results.append(fut.result())
            except Exception as e:
                sym = futs[fut]
                log.warning("%s failed: %s", sym, e)
                results.append({"Symbol": sym, "Data Status": f"ERROR: {e}"})
            if i % 50 == 0:
                log.info("Processed %d/%d", i, len(futs))

    df = pd.DataFrame(results)
    if df.empty:
        raise RuntimeError("No results generated.")

    score_order = {"3/3": 0, "2/3": 1, "1/3": 2, "0/3": 3}
    df["_sort"] = df["Score"].map(score_order).fillna(9)
    df = df.sort_values(["_sort", "52W Return"], ascending=[True, False]).drop(columns="_sort")

    date = datetime.now().strftime("%Y-%m-%d")
    path = OUT / f"Mehta_Screener_{date}.xlsx"
    export_excel(df, path)

    log.info("Done: %s", path)
    log.info("Elapsed: %.1fs", time.time() - start)

    print("\nMEHTA 3/3 SCREEN COMPLETE")
    print(f"Stocks screened: {len(df)}")
    print(f"3/3: {(df['Score'] == '3/3').sum()}")
    print(f"2/3: {(df['Score'] == '2/3').sum()}")
    print(f"1/3: {(df['Score'] == '1/3').sum()}")
    print(f"0/3: {(df['Score'] == '0/3').sum()}")
    print(f"Excel: {path}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("Interrupted.")
        sys.exit(130)
    except Exception as e:
        log.exception("Fatal error: %s", e)
        sys.exit(1)
